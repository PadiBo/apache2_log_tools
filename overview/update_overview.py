import paramiko
import time
import os
import gzip
import re
import pandas as pd
import shutil
from datetime import datetime
from openpyxl import load_workbook

def download_files(server_ip, username, password):
    # connect to server
    ssh = paramiko.SSHClient()
    ssh.set_missing_host_key_policy(paramiko.AutoAddPolicy())
    ssh.connect(server_ip, username=username, password=password)

    # SFTP-Connection
    sftp = ssh.open_sftp()

    # local path for downloaded files
    local_directory = "./downloaded_files"
    if not os.path.exists(local_directory):
        os.makedirs(local_directory)

    # download, decompress and combine logs to a single file
    remote_directory = "/var/log/apache2/"
    for filename in sftp.listdir(remote_directory):
        if filename.startswith("access"):  # check for files starting with "access"
            remote_file_path = os.path.join(remote_directory, filename)
            local_file_path = os.path.join(local_directory, filename)
            sftp.get(remote_file_path, local_file_path)

            # unzip file if its compressed
            if filename.endswith(".gz"):
                with gzip.open(local_file_path, 'rb') as f_in:
                    with open(local_file_path[:-3], 'wb') as f_out:
                        f_out.write(f_in.read())
                os.remove(local_file_path)  # delete .gz files

    # close sftp and ssh
    sftp.close()
    ssh.close()

    print("1/6 all access logs downloaded and decompressed.")
    return local_directory

from datetime import datetime, timedelta

def parse_apache_log(log_file):
    log_pattern = r'(?P<remote_ip>\d+\.\d+\.\d+\.\d+) - (?P<user>\S+) \[(?P<timestamp>.*?)\] "(?P<method>\w+) (?P<url>.*?) \S+" (?P<status>\d+) (?P<bytes_sent>\d+) "(?P<referrer>.*?)" "(?P<user_agent>.*?)"'

    data = []

    # load ips to exclude from overview.xlsx sheet excludeIP
    exclude_ip_df = pd.read_excel('overview.xlsx', sheet_name='excludeIP')
    excluded_ips = exclude_ip_df['IP'].tolist()

    with open(log_file, 'r') as file:
        for line in file:
            match = re.match(log_pattern, line)
            if match:
                log_entry = match.groupdict()
                remote_ip = log_entry['remote_ip']
                # check if ip should be excluded (for local ips and your own)
                if remote_ip not in excluded_ips:
                    # split time and date
                    timestamp_str = log_entry['timestamp']
                    timestamp_obj = datetime.strptime(timestamp_str, '%d/%b/%Y:%H:%M:%S %z')
                    log_entry['date'] = timestamp_obj.strftime('%d.%m.%Y')
                    log_entry['time'] = timestamp_obj.strftime('%H:%M:%S')
                    
                  # HTTP-Status-Codes
                    status_code = log_entry['status']
                    status_description = get_status_description(status_code)
                    log_entry['status_description'] = status_description
                    data.append(log_entry)

    # sort by date and time, newest on top
    sorted_data = sorted(data, key=lambda x: datetime.strptime(x['date'] + ' ' + x['time'], '%d.%m.%Y %H:%M:%S'), reverse=True)

    return sorted_data

def get_status_description(status_code):
    status_codes = {
        '100': 'Continue',
        '101': 'Switching Protocols',
        '200': 'OK',
        '201': 'Created',
        '202': 'Accepted',
        '203': 'Non-Authoritative Information',
        '204': 'No Content',
        '205': 'Reset Content',
        '206': 'Partial Content',
        '300': 'Multiple Choices',
        '301': 'Moved Permanently',
        '302': 'Found',
        '303': 'See Other',
        '304': 'Not Modified',
        '305': 'Use Proxy',
        '307': 'Temporary Redirect',
        '308': 'Permanent Redirect',
        '400': 'Bad Request',
        '401': 'Unauthorized',
        '402': 'Payment Required',
        '403': 'Forbidden',
        '404': 'Not Found',
        '405': 'Method Not Allowed',
        '406': 'Not Acceptable',
        '407': 'Proxy Authentication Required',
        '408': 'Request Timeout',
        '409': 'Conflict',
        '410': 'Gone',
        '411': 'Length Required',
        '412': 'Precondition Failed',
        '413': 'Payload Too Large',
        '414': 'URI Too Long',
        '415': 'Unsupported Media Type',
        '416': 'Range Not Satisfiable',
        '417': 'Expectation Failed',
        '418': "I'm a teapot",
        '421': 'Misdirected Request',
        '422': 'Unprocessable Entity',
        '423': 'Locked',
        '424': 'Failed Dependency',
        '425': 'Too Early',
        '426': 'Upgrade Required',
        '428': 'Precondition Required',
        '429': 'Too Many Requests',
        '431': 'Request Header Fields Too Large',
        '451': 'Unavailable For Legal Reasons',
        '500': 'Internal Server Error',
        '501': 'Not Implemented',
        '502': 'Bad Gateway',
        '503': 'Service Unavailable',
        '504': 'Gateway Timeout',
        '505': 'HTTP Version Not Supported',
        '506': 'Variant Also Negotiates',
        '507': 'Insufficient Storage',
        '508': 'Loop Detected',
        '510': 'Not Extended',
        '511': 'Network Authentication Required'
    }
    return status_codes.get(status_code, 'Unknown')

def write_to_excel(data, output_file):
    df = pd.DataFrame(data, columns=['remote_ip', 'user', 'date', 'time', 'method', 'url', 'status', 'status_description', 'bytes_sent', 'referrer', 'user_agent'])
    df.to_excel(output_file, index=False)
    print("2/6 data sucessfully written to {}.".format(output_file))

def update_request_sheet(access_file, overview_file):
    # read data from overview and access file
    print("3/6 read data.")
    access_df = pd.read_excel(access_file)
    try:
        overview_df = pd.read_excel(overview_file, sheet_name='rqst')
    except:
        overview_df = pd.DataFrame()

    # add new data
    print("4/6 combine data.")
    # combine current data to existing data
    updated_requests_df = pd.concat([overview_df, access_df], ignore_index=True)
    
    # remove duplicates
    print("5/6 remove duplicates.")
    updated_requests_df.drop_duplicates(inplace=True)
    
    # sort data 
    print("Sorting data.")
    updated_requests_df['datetime'] = pd.to_datetime(updated_requests_df['date'] + ' ' + updated_requests_df['time'], dayfirst=True)
    updated_requests_df.sort_values(by='datetime', ascending=False, inplace=True)
    updated_requests_df.drop(columns='datetime', inplace=True)

    # write to rqst sheet 
    print("6/6 Data written to excelfile.")
    with pd.ExcelWriter(overview_file, engine="openpyxl", mode="a") as writer:
        if 'rqst' in writer.book.sheetnames:
            idx = writer.book.sheetnames.index('rqst')
            writer.book.remove(writer.book.worksheets[idx])

        updated_requests_df.to_excel(writer, index=False, sheet_name='rqst')

    print("Done.")
    
    # delete acces.xlsx
    os.remove(access_file)
    print("The file {} was deleted.".format(access_file))

    # delete downloaded_files folder
    shutil.rmtree("./downloaded_files")
    print("The folder 'downloaded_files' was deleted.")
    input("Press Enter to exit.")
    print("Bye!")
    time.sleep(1)

# Script call
server_ip = "serveriphere" #ADD SERVERIP
username = "usernamehere"  #ADD USERNAME
password = "passwordhere"  #ADD PASSWORD

local_directory = download_files(server_ip, username, password)
parsed_data = []
for filename in os.listdir(local_directory):
    if filename.startswith("access"):
        file_path = os.path.join(local_directory, filename)
        parsed_data.extend(parse_apache_log(file_path))

output_excel_file = "access_logs.xlsx"
write_to_excel(parsed_data, output_excel_file)

access_file = "access_logs.xlsx"
overview_file = "overview.xlsx"
update_request_sheet(access_file, overview_file)
